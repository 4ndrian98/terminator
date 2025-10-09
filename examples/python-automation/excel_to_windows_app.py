"""
Automation example: read rows from an Excel sheet and populate a Windows form.

Usage notes:
- Install dependencies: `pip install openpyxl`.
- Update FIELD_MAPPING, WINDOW_SELECTOR, and button/selectors to match your application.
- Optionally provide environment variables (EXCEL_FILE, EXCEL_SHEET, TARGET_APP, APP_ARGS).
"""

import os
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency 'openpyxl'. Install it with `pip install openpyxl`."
    ) from exc


EXCEL_FILE = os.environ.get("EXCEL_FILE", r"C:\path\to\spreadsheet.xlsx")
EXCEL_SHEET = os.environ.get("EXCEL_SHEET", "Sheet1")

# Map Excel column headers to UI element selectors. Update to match your form.
FIELD_MAPPING: Dict[str, str] = {
    "Nome": "role:Edit|name:Nome",
    "Cognome": "role:Edit|name:Cognome",
    "Email": "role:Edit|name:Email",
}

TARGET_APP = os.environ.get("TARGET_APP")  # e.g. r"C:\Program Files\MyApp\myapp.exe"
APP_ARGS = os.environ.get("APP_ARGS")  # e.g. "--profile default"

WINDOW_SELECTOR = os.environ.get(
    "APP_WINDOW_SELECTOR", "role:Window|name:My Desktop App"
)

SAVE_BUTTON_SELECTOR = os.environ.get("SAVE_BUTTON_SELECTOR", "role:Button|name:Salva")
NEXT_RECORD_SHORTCUT = os.environ.get("NEXT_RECORD_SHORTCUT", "{Alt+n}")

LAUNCH_WAIT_MS = int(os.environ.get("LAUNCH_WAIT_MS", "4000"))
ACTION_DELAY_MS = int(os.environ.get("ACTION_DELAY_MS", "300"))
BETWEEN_ROWS_DELAY_MS = int(os.environ.get("BETWEEN_ROWS_DELAY_MS", "700"))


def _read_excel(path: str, sheet_name: str) -> List[Dict[str, str]]:
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in '{path}'.")
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    rows: List[Dict[str, str]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        row = {}
        for header, cell in zip(headers, raw_row):
            if not header:
                continue
            value = "" if cell is None else str(cell).strip()
            row[str(header).strip()] = value
        if any(value for value in row.values()):
            rows.append(row)
    return rows


async def _wait_for_window(selector: str, attempts: int = 8) -> Optional[object]:
    for attempt in range(1, attempts + 1):
        window = await desktop.locator(selector).first()
        if window:
            return window
        log(f"Window not found, retry {attempt}/{attempts}...")
        await sleep(750)
    return None


async def _fill_field(selector: str, value: str) -> bool:
    element = await desktop.locator(selector).first()
    if not element:
        log(f"Selector '{selector}' not found.")
        return False
    element.click()
    await sleep(ACTION_DELAY_MS)
    element.press_key("{Control+a}")
    await sleep(ACTION_DELAY_MS // 2)
    if value:
        element.type_text(value)
    else:
        element.type_text("")
    await sleep(ACTION_DELAY_MS)
    return True


log("Excel → Windows form automation starting.")

excel_path = Path(EXCEL_FILE)
if not excel_path.exists():
    log(f"Excel file not found at {excel_path}.")
    return {
        "status": "error",
        "message": f"Excel file not found at {excel_path}",
    }

try:
    excel_rows = _read_excel(str(excel_path), EXCEL_SHEET)
except Exception as error:
    log(f"Failed to read Excel data: {error}")
    return {"status": "error", "message": f"Excel read failed: {error}"}

if not excel_rows:
    log("No data rows detected in the Excel sheet.")
    return {"status": "error", "message": "Excel sheet has no data rows."}

log(f"Loaded {len(excel_rows)} rows from Excel.")

if TARGET_APP:
    command = TARGET_APP if not APP_ARGS else f"{TARGET_APP} {APP_ARGS}"
    log(f"Launching target application: {command}")
    desktop.open_application(command)
    await sleep(LAUNCH_WAIT_MS)

window = await _wait_for_window(WINDOW_SELECTOR)
if not window:
    log("Could not attach to the target window.")
    return {
        "status": "error",
        "message": f"Window '{WINDOW_SELECTOR}' not found.",
    }

processed_rows = []
failed_rows = []

for idx, row in enumerate(excel_rows, start=1):
    log(f"Processing Excel row #{idx}")
    row_success = True
    for header, selector in FIELD_MAPPING.items():
        value = row.get(header, "")
        filled = await _fill_field(selector, value)
        if not filled:
            row_success = False
    if row_success and SAVE_BUTTON_SELECTOR:
        save_button = await desktop.locator(SAVE_BUTTON_SELECTOR).first()
        if save_button:
            save_button.click()
            log("Save button clicked.")
            await sleep(ACTION_DELAY_MS)
        else:
            log(f"Save button '{SAVE_BUTTON_SELECTOR}' not found.")
            row_success = False
    if row_success:
        processed_rows.append(idx)
    else:
        failed_rows.append(idx)
    if NEXT_RECORD_SHORTCUT and idx != len(excel_rows):
        log(f"Advancing to next record with shortcut {NEXT_RECORD_SHORTCUT}.")
        desktop.press_key(NEXT_RECORD_SHORTCUT)
        await sleep(BETWEEN_ROWS_DELAY_MS)

log("Automation finished.")

return {
    "status": "success" if not failed_rows else "partial",
    "processed_rows": processed_rows,
    "failed_rows": failed_rows,
    "total_rows": len(excel_rows),
    "message": "Completed form population from Excel.",
}
